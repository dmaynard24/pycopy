from openpyxl import load_workbook
import os, html

# default config, will probably need to tweak a couple of these before running
default_categories = 'CCFE,APP'
default_language = 'en_US'
default_protected = 'false'
xl_dir = 'C:\\Users\\dmayna\\Desktop\\CareerChoice\\Stories\\US8309\\translations\\'
base_out_dir = 'C:\\repos\\ccnextgen-sfdx\\force-app\\main\\default\\'
# always open the "master", this will hold custom label names
master_wb = load_workbook(
	filename=f'{xl_dir}App Flow Content FINAL - ENG.xlsx')
filenames = {
	'fr_CA': 'App Flow Content FINAL, FR - CA',
	'cs': 'App Flow Content FINAL, CZ',
	'de': 'App Flow Content FINAL, DE',
	'es': 'App Flow Content FINAL, ES',
	'fr': 'App Flow Content FINAL, FR -FR',
	'it': 'App Flow Content FINAL, IT',
	'pl': 'App Flow Content FINAL, POL',
	'sk': 'App Flow Content FINAL, SK',
}
sheet_name = '08 Review'
wb_sheet_master = master_wb[sheet_name]
row_count = wb_sheet_master.max_row  # last row with content


def write_xml(out_dir, filename, xml, root_node):
	prefix = f'<?xml version="1.0" encoding="UTF-8"?>\n<{root_node} xmlns="http://soap.sforce.com/2006/04/metadata">\n'
	postfix = f'\n</{root_node}>'
	# concat entire file contents
	xml = f'{prefix}{xml.rstrip()}{postfix}'
	xml_file = f'{out_dir}{filename}'
	try:
		os.mkdir(out_dir)
	except FileExistsError:
		pass
	with open(xml_file, 'w', encoding='utf-8') as f:
		f.write(xml)


def write_custom_labels():
	print('Writing custom labels.')
	# read master workbook and create XML nodes for each label
	xml = ''
	for row in range(1, row_count + 1):
		label = wb_sheet_master[f'A{row}'].value
		name = wb_sheet_master[f'B{row}'].value
		categories = wb_sheet_master[f'C{row}'].value
		if label is None or label.strip() == '' or name is None or name.strip(
		) == '':
			# empty cell
			continue
		label = html.escape(label)
		xml += f'\t<labels>\n\t\t<fullName>{name}</fullName>'
		xml += f'\n\t\t<categories>{default_categories if categories is None else categories}</categories>'
		xml += f'\n\t\t<language>{default_language}</language>'
		xml += f'\n\t\t<protected>{default_protected}</protected>'
		xml += f'\n\t\t<shortDescription>{name}</shortDescription>'
		xml += f'\n\t\t<value>{label}</value>'
		xml += f'\n\t</labels>\n'

	# write the concatenated xml to a file
	out_dir = f'{base_out_dir}labels\\'
	filename = 'CustomLabels.labels-meta.xml'
	write_xml(out_dir, filename, xml, 'CustomLabels')


def write_translations():
	# read all translation workbooks
	for lang in filenames.keys():
		print(f'{lang}: Writing translations.')
		wb_filename = filenames[lang]
		wb = load_workbook(filename=f'{xl_dir}{wb_filename}.xlsx')
		wb_sheet = wb[sheet_name]

		# catch error if this workbook has different number of rows, the data may need to be massaged
		wb_sheet_row_count = wb_sheet.max_row
		if wb_sheet_row_count != row_count:
			print(
				f'Error occurred. The {lang} workbook has {wb_sheet_row_count} while the master workbook has {row_count}. This may not affect the output negatively, sometimes max_row is unpredictable.'
			)

		# go over all rows and create XML nodes for each lang
		xml = ''
		for row in range(1, row_count + 1):
			label = wb_sheet[f'A{row}'].value
			name = wb_sheet_master[f'B{row}'].value
			if label is None or label.strip() == '' or name is None or name.strip(
			) == '':
				# empty cell
				continue
			label = html.escape(label)
			xml += f'\t<customLabels>\n\t\t<label>{label}</label>\n\t\t<name>{name}</name>\n\t</customLabels>\n'

		# write the concatenated xml to a file
		out_dir = f'{base_out_dir}translations\\'
		filename = f'{lang}.translation-meta.xml'
		write_xml(out_dir, filename, xml, 'Translations')


write_custom_labels()
write_translations()
print('Done!')